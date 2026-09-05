# Minimal runnable fixtures for the snippets called out in review: one file per format.
# Each script is self-contained, writes only scratch files into the current directory,
# and exits non-zero on failed assertions. Run from any scratch directory:
#   python pdf_fixture.py   (deps: reportlab, pypdf, pymupdf)
#   python pptx_fixture.py  (deps: python-pptx)
#   python xlsx_fixture.py  (deps: openpyxl)
#   python docx_fixture.py  (deps: python-docx, pymupdf, soffice on PATH)
import base64
import csv
import os
import sys
import zipfile
from pathlib import Path
from tempfile import TemporaryFile, mkstemp

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

failures = []


def check(name, cond, extra=""):
    print(("PASS " if cond else "FAIL ") + name + ((" :: " + str(extra)) if not cond and extra else ""))
    if not cond:
        failures.append(name)


# Execute the canonical package.md snippet itself so fixtures cannot drift from the Skill.
package_reference = (
    Path(__file__).resolve().parents[1] / "skills" / "document-skills-xlsx" / "references" / "package.md"
)
package_markdown = package_reference.read_text(encoding="utf-8")
package_code = package_markdown.split("```python", 1)[1].split("```", 1)[0]
exec(compile(package_code, str(package_reference), "exec"), globals())


# ---- csv.md snippet: sniffed dialect actually reaches the reader ---------------
semicolon_csv = "region;units;note\nEU;120;first\nUS;80;second\n"
with open("input.csv", "w", newline="", encoding="utf-8") as f:
    f.write(semicolon_csv)

with open("input.csv", newline="", encoding="utf-8-sig") as f:
    sample = f.read(2048)
    f.seek(0)
    dialect = csv.Sniffer().sniff(sample)
    reader = csv.DictReader(f, dialect=dialect)
    rows = list(reader)

check("sniffer detects the semicolon dialect", dialect.delimiter == ";", dialect.delimiter)
check("rows parse into per-column values", rows[0] == {"region": "EU", "units": "120", "note": "first"}, rows[0])

# negative control: default comma reader splits the whole line as one key
with open("input.csv", newline="", encoding="utf-8-sig") as f:
    naive = next(csv.DictReader(f))
check("default reader is proven wrong here (negative control)", list(naive.keys())[0] == "region;units;note", naive)

# ---- csv.md conversion: formula-looking input remains literal text -----------
formula_looking = '=HYPERLINK("https://example.invalid", "click")'
csv_wb = openpyxl.Workbook()
csv_cell = csv_wb.active["A1"]
csv_cell.value = formula_looking
csv_cell.data_type = "s"
csv_wb.save("csv-text.xlsx")
csv_reopened = openpyxl.load_workbook("csv-text.xlsx", data_only=False)
check("formula-looking CSV field keeps its exact text", csv_reopened.active["A1"].value == formula_looking)
check("formula-looking CSV field is not an XLSX formula", csv_reopened.active["A1"].data_type == "s")

unsafe_wb = openpyxl.Workbook()
unsafe_wb.active["A1"] = formula_looking
check("plain assignment is proven unsafe (negative control)", unsafe_wb.active["A1"].data_type == "f")

CSV_FORMULA_OPERATORS = ("=", "+", "-", "@", "＝", "＋", "－", "＠")


def begins_spreadsheet_formula(value):
    index = 0
    while index < len(value) and (ord(value[index]) <= 0x20 or value[index] == "\ufeff"):
        index += 1
    return value.startswith(CSV_FORMULA_OPERATORS, index)


def spreadsheet_csv_field(value, *, mode="safe"):
    if mode not in {"safe", "raw"}:
        raise ValueError("mode must be 'safe' or 'raw'")
    if mode == "safe" and isinstance(value, str) and begins_spreadsheet_formula(value):
        return "'" + value
    return value


def delimiter_for(path):
    delimiter = {".csv": ",", ".tsv": "\t"}.get(Path(path).suffix.lower())
    if delimiter is None:
        raise ValueError("output must use a .csv or .tsv extension")
    return delimiter


formula_like_fields = [
    "\ufeff=1+1", "=1+1", "+SUM(A1:A2)", "-2+3", "@cmd",
    "\t=1+1", "\r@cmd", "\n-2+3", "\x00\t +SUM(A1:A2)",
    "＝1+1", "＋SUM(A1:A2)", "－2+3", "＠cmd", "-7",
]
benign_fields = ["plain", -7, "\tplain", "\rplain", "\nplain", "\ufeffplain", "\x00 plain"]
with open("spreadsheet-safe.csv", "w", newline="", encoding="utf-8") as output:
    csv.writer(output).writerow([
        spreadsheet_csv_field(value) for value in formula_like_fields + benign_fields
    ])
with open("spreadsheet-safe.csv", newline="", encoding="utf-8-sig") as exported:
    safe_fields = next(csv.reader(exported))
check("spreadsheet-safe CSV neutralizes operator, control, BOM, and fullwidth prefixes",
      safe_fields[:len(formula_like_fields)]
      == ["'" + value for value in formula_like_fields], safe_fields)
check("safe CSV preserves benign text and numeric values",
      safe_fields[len(formula_like_fields):]
      == [str(value) for value in benign_fields], safe_fields)
check("raw CSV mode preserves exact formula-like literal strings",
      [spreadsheet_csv_field(value, mode="raw") for value in formula_like_fields]
      == formula_like_fields)
try:
    spreadsheet_csv_field("=1+1", mode="unknown")
    invalid_csv_mode_rejected = False
except ValueError:
    invalid_csv_mode_rejected = True
check("CSV export rejects an ambiguous safety mode", invalid_csv_mode_rejected)

tabular_rows = [["Region", "Units", "Note"], ["EU, West", 120, "\t=1+1"]]
safe_tabular_rows = [
    [spreadsheet_csv_field(value) for value in row] for row in tabular_rows
]
with open("spreadsheet-safe.TSV", "w", newline="", encoding="utf-8") as output:
    csv.writer(output, delimiter=delimiter_for("spreadsheet-safe.TSV")).writerows(safe_tabular_rows)
with open("spreadsheet-safe.TSV", newline="", encoding="utf-8") as exported:
    tsv_rows = list(csv.reader(exported, delimiter="\t"))
check("TSV export selects a tab delimiter case-insensitively",
      tsv_rows == [[str(value) for value in row] for row in safe_tabular_rows], tsv_rows)
check("TSV safe mode neutralizes a control-prefixed formula after round-trip",
      tsv_rows[1][2] == "'\t=1+1", tsv_rows)
check("CSV export retains its comma delimiter", delimiter_for("output.csv") == ",")
try:
    delimiter_for("output.txt")
    unknown_tabular_suffix_rejected = False
except ValueError:
    unknown_tabular_suffix_rejected = True
check("tabular export rejects an unknown extension", unknown_tabular_suffix_rejected)

# ---- edit.md snippet: round_trip_changes detects dropped parts AND stripped extensions ----
from collections import Counter
from xml.etree import ElementTree as ET

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Data"
ws.append(["Region", "Units"])
ws.append(["EU", 120])
wb.create_sheet("Keep")["A1"] = "keep"
wb.save("plain.xlsx")

with zipfile.ZipFile("plain.xlsx") as zin:
    payload = {name: zin.read(name) for name in zin.namelist()}

# ---- package.md: preflight runs before openpyxl and bounds every package part -----
try:
    with open_validated_workbook("plain.xlsx") as validated_plain:
        ordinary_package_loaded = validated_plain.sheetnames == ["Data", "Keep"]
except Exception as error:
    ordinary_package_loaded = False
    ordinary_package_error = error
check("bounded XLSX loader accepts an ordinary package", ordinary_package_loaded,
      locals().get("ordinary_package_error"))

compressed_payload = dict(payload)
compressed_payload["xl/styles.xml"] = (
    b'<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
    + b"A" * (1024 * 1024)
    + b"</styleSheet>"
)
with zipfile.ZipFile("compressed-xlsx-bomb.xlsx", "w", zipfile.ZIP_DEFLATED) as archive:
    for name, data in compressed_payload.items():
        archive.writestr(name, data)
real_load_workbook = openpyxl.load_workbook
preflight_load_calls = []
openpyxl.load_workbook = lambda *args, **kwargs: preflight_load_calls.append(args) or None
try:
    with open_validated_workbook("compressed-xlsx-bomb.xlsx"):
        compressed_bomb_rejected = False
except ValueError as error:
    compressed_bomb_rejected = "suspicious compression ratio: xl/styles.xml" in str(error)
finally:
    openpyxl.load_workbook = real_load_workbook
check("XLSX compression bomb is rejected before openpyxl runs",
      compressed_bomb_rejected and preflight_load_calls == [],
      (compressed_bomb_rejected, preflight_load_calls))
check("XLSX preflight gates remain active under optimized Python",
      __debug__ or compressed_bomb_rejected)

original_member_limit = MAX_MEMBERS
try:
    with zipfile.ZipFile("plain.xlsx") as archive:
        MAX_MEMBERS = len(archive.infolist()) - 1
    with validated_xlsx_source("plain.xlsx"):
        member_limit_rejected = False
except ValueError as error:
    member_limit_rejected = str(error) == "archive member count above limit"
finally:
    MAX_MEMBERS = original_member_limit
check("XLSX member-count limit rejects a real package before parsing", member_limit_rejected)

metadata_names = [
    name for name in payload
    if name.endswith((".xml", ".rels")) and not name.startswith("xl/worksheets/")
]
metadata_test_limit = max(len(payload[name]) for name in metadata_names) + 128
oversized_metadata = dict(payload)
oversized_metadata["xl/styles.xml"] = (
    b'<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><!--'
    + b"x" * metadata_test_limit
    + b"--></styleSheet>"
)
with zipfile.ZipFile("oversized-metadata.xlsx", "w", zipfile.ZIP_STORED) as archive:
    for name, data in oversized_metadata.items():
        archive.writestr(name, data)
original_xml_limit = MAX_XML_PART
try:
    MAX_XML_PART = metadata_test_limit
    with validated_xlsx_source("oversized-metadata.xlsx"):
        oversized_metadata_rejected = False
except ValueError as error:
    oversized_metadata_rejected = str(error) == "oversized XML part: xl/styles.xml"
finally:
    MAX_XML_PART = original_xml_limit
check("metadata XML part limit is enforced before openpyxl", oversized_metadata_rejected)

large_worksheet_payload = dict(payload)
sheet_name = "xl/worksheets/sheet1.xml"
large_worksheet_payload[sheet_name] = large_worksheet_payload[sheet_name].replace(
    b"</worksheet>", b"<!--" + b"w" * (metadata_test_limit + 256) + b"--></worksheet>"
)
with zipfile.ZipFile("large-streamed-worksheet.xlsx", "w", zipfile.ZIP_STORED) as archive:
    for name, data in large_worksheet_payload.items():
        archive.writestr(name, data)
original_xml_limit = MAX_XML_PART
original_worksheet_limit = MAX_WORKSHEET_XML
try:
    MAX_XML_PART = metadata_test_limit
    MAX_WORKSHEET_XML = len(large_worksheet_payload[sheet_name]) + 128
    with validated_xlsx_source("large-streamed-worksheet.xlsx"):
        streamed_worksheet_passed = True
except Exception as error:
    streamed_worksheet_passed = False
    streamed_worksheet_error = error
finally:
    MAX_XML_PART = original_xml_limit
    MAX_WORKSHEET_XML = original_worksheet_limit
check("large worksheet XML uses the bounded streaming limit, not metadata limit",
      streamed_worksheet_passed, locals().get("streamed_worksheet_error"))

unsafe_xml_payload = dict(payload)
unsafe_xml_payload["xl/unsafe.xml"] = (
    '<?xml version="1.0" encoding="UTF-16"?>'
    '<!DOCTYPE unsafe [<!ENTITY x "expanded">]><unsafe>&x;</unsafe>'
).encode("utf-16")
with zipfile.ZipFile("unsafe-utf16-xml.xlsx", "w", zipfile.ZIP_STORED) as archive:
    for name, data in unsafe_xml_payload.items():
        archive.writestr(name, data)
try:
    with validated_xlsx_source("unsafe-utf16-xml.xlsx"):
        utf16_dtd_rejected = False
except ValueError as error:
    utf16_dtd_rejected = "unsafe or malformed XML part: xl/unsafe.xml" in str(error)
check("defused XML parsing rejects UTF-16 DTD/entity parts", utf16_dtd_rejected)

semantic_range_payloads = {
    "merge-range-bomb.xlsx": (
        b'<mergeCells count="1"><mergeCell ref="A1:XFD1048576"/></mergeCells>'
    ),
    "hyperlink-range-bomb.xlsx": (
        b'<hyperlinks><hyperlink ref="A1:XFD1048576" location="Data!A1"/>'
        b'</hyperlinks>'
    ),
}
for malicious_path, range_markup in semantic_range_payloads.items():
    malicious_payload = dict(payload)
    malicious_payload[sheet_name] = malicious_payload[sheet_name].replace(
        b"</worksheet>", range_markup + b"</worksheet>"
    )
    with zipfile.ZipFile(malicious_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, data in malicious_payload.items():
            archive.writestr(name, data)

from openpyxl.comments import Comment

comment_wb = openpyxl.Workbook()
comment_wb.active["A1"].comment = Comment("bounded note", "fixture")
comment_wb.save("comment-base.xlsx")
comment_wb.close()
with zipfile.ZipFile("comment-base.xlsx") as archive:
    comment_payload = {name: archive.read(name) for name in archive.namelist()}
comment_part = next(
    name for name in comment_payload
    if name.casefold().startswith("xl/comments/") and name.casefold().endswith(".xml")
)
comment_payload[comment_part] = comment_payload[comment_part].replace(
    b'ref="A1"', b'ref="A1:XFD1048576"', 1
)
with zipfile.ZipFile("comment-range-bomb.xlsx", "w", zipfile.ZIP_DEFLATED) as archive:
    for name, data in comment_payload.items():
        archive.writestr(name, data)

semantic_expected_errors = {
    **{path: "worksheet cell materialization budget exceeded"
       for path in semantic_range_payloads},
    "comment-range-bomb.xlsx": "comment reference must identify one cell",
}

real_load_workbook = openpyxl.load_workbook
semantic_preflight_load_calls = []
semantic_range_results = {}
openpyxl.load_workbook = (
    lambda *args, **kwargs: semantic_preflight_load_calls.append(args) or None
)
try:
    for malicious_path, expected_error in semantic_expected_errors.items():
        try:
            with open_validated_workbook(malicious_path):
                semantic_range_results[malicious_path] = False
        except ValueError as error:
            semantic_range_results[malicious_path] = (
                expected_error in str(error)
            )

    original_materialized_cell_limit = MAX_WORKSHEET_MATERIALIZED_CELLS
    try:
        MAX_WORKSHEET_MATERIALIZED_CELLS = 1
        with open_validated_workbook("plain.xlsx"):
            explicit_cell_limit_rejected = False
    except ValueError as error:
        explicit_cell_limit_rejected = (
            str(error) == "worksheet cell materialization budget exceeded"
        )
    finally:
        MAX_WORKSHEET_MATERIALIZED_CELLS = original_materialized_cell_limit
finally:
    openpyxl.load_workbook = real_load_workbook

check(
    "merge, hyperlink, and comment ranges are bounded before openpyxl runs",
    all(semantic_range_results.values()) and semantic_preflight_load_calls == [],
    (semantic_range_results, semantic_preflight_load_calls),
)
check(
    "explicit worksheet cells share the pre-load materialization budget",
    explicit_cell_limit_rejected and semantic_preflight_load_calls == [],
    (explicit_cell_limit_rejected, semantic_preflight_load_calls),
)
with open_validated_workbook("comment-base.xlsx") as validated_comment_wb:
    valid_comment_loaded = (
        validated_comment_wb.active["A1"].comment is not None
        and validated_comment_wb.active["A1"].comment.text == "bounded note"
    )
check("a valid single-cell comment survives bounded loading", valid_comment_loaded)

range_list_payloads = {
    "range-list-budget.xlsx": b'<conditionalFormatting sqref="A1 B1 C1"/>',
    "range-list-whole-sheet.xlsx": (
        b'<conditionalFormatting sqref="A1:XFD1048576"/>'
    ),
}
for range_list_path, range_list_markup in range_list_payloads.items():
    range_list_payload = dict(payload)
    range_list_payload[sheet_name] = range_list_payload[sheet_name].replace(
        b"</worksheet>", range_list_markup + b"</worksheet>"
    )
    with zipfile.ZipFile(range_list_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, data in range_list_payload.items():
            archive.writestr(name, data)

original_range_token_limit = MAX_WORKSHEET_RANGE_TOKENS
original_range_list_char_limit = MAX_WORKSHEET_RANGE_LIST_CHARS
real_load_workbook = openpyxl.load_workbook
range_list_load_calls = []
openpyxl.load_workbook = lambda *args, **kwargs: range_list_load_calls.append(args) or None
try:
    MAX_WORKSHEET_RANGE_TOKENS = 2
    try:
        with open_validated_workbook("range-list-budget.xlsx"):
            range_token_budget_rejected = False
    except ValueError as error:
        range_token_budget_rejected = str(error) == "worksheet range-token budget exceeded"
    MAX_WORKSHEET_RANGE_TOKENS = 1
    try:
        with validated_xlsx_source("range-list-whole-sheet.xlsx"):
            whole_sheet_range_accepted = True
    except Exception as error:
        whole_sheet_range_accepted = False
        whole_sheet_range_error = error
    MAX_WORKSHEET_RANGE_LIST_CHARS = 8
    try:
        with validated_xlsx_source("range-list-whole-sheet.xlsx"):
            long_range_list_rejected = False
    except ValueError as error:
        long_range_list_rejected = "worksheet range list is too long" in str(error)
finally:
    MAX_WORKSHEET_RANGE_TOKENS = original_range_token_limit
    MAX_WORKSHEET_RANGE_LIST_CHARS = original_range_list_char_limit
    openpyxl.load_workbook = real_load_workbook

check(
    "worksheet range-list token growth is bounded before openpyxl runs",
    range_token_budget_rejected and range_list_load_calls == [],
    (range_token_budget_rejected, range_list_load_calls),
)
check(
    "one whole-sheet sqref costs one range token rather than its cell area",
    whole_sheet_range_accepted,
    locals().get("whole_sheet_range_error"),
)
check("worksheet sqref character growth is bounded", long_range_list_rejected)

Path("snapshot-source.xlsx").write_bytes(Path("plain.xlsx").read_bytes())
replacement_wb = openpyxl.Workbook()
replacement_wb.active["A1"] = "UNVALIDATED"
replacement_wb.save("snapshot-replacement.xlsx")
replacement_wb.close()
with validated_xlsx_source("snapshot-source.xlsx") as snapshot_source:
    Path("snapshot-source.xlsx").write_bytes(
        Path("snapshot-replacement.xlsx").read_bytes()
    )
    snapshot_wb = openpyxl.load_workbook(snapshot_source, read_only=True, data_only=True)
    try:
        snapshot_value = snapshot_wb.active["A1"].value
    finally:
        snapshot_wb.close()
check(
    "validated source is a private snapshot isolated from later path changes",
    snapshot_value == "Region",
    snapshot_value,
)

# simulate an unsupported extension part (what a slicer/queries part looks like in the zip)
payload["xl/slicers/slicer1.xml"] = b"<slicer xmlns='stub'/>"
payload["xl/media/large.bin"] = b"x14:" * 32_768  # binary payload must never be marker-scanned

# Two records deliberately share a URI/namespace but carry different child content. A coarse
# presence set cannot see one disappear while the other survives.
SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
X14_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
EXT_URI = "{00000000-0000-0000-0000-000000000000}"
KEEP_EXT = (
    f'<ext uri="{EXT_URI}" xmlns:x14="{X14_NS}"><x14:stub id="keep"/></ext>'
).encode()
KEEP_EXT_ALT_PREFIX = (
    f'<ext uri="{EXT_URI}" xmlns:alt="{X14_NS}"><alt:stub id="keep"/></ext>'
).encode()
DROP_EXT = (
    f'<ext uri="{EXT_URI}" xmlns:sx="{X14_NS}"><sx:stub id="drop"/></ext>'
).encode()
EXT_LIST_BOTH = b"<extLst>" + KEEP_EXT + DROP_EXT + b"</extLst>"
EXT_LIST_KEEP = b"<extLst>" + KEEP_EXT_ALT_PREFIX + b"</extLst>"
EXT_LIST_DROP = b"<extLst>" + DROP_EXT + b"</extLst>"
payload["xl/worksheets/sheet1.xml"] = payload["xl/worksheets/sheet1.xml"].replace(
    b"</worksheet>", EXT_LIST_BOTH + b"</worksheet>")
payload["xl/worksheets/sheet2.xml"] = payload["xl/worksheets/sheet2.xml"].replace(
    b"</worksheet>", EXT_LIST_DROP + b"</worksheet>")

with zipfile.ZipFile("extended.xlsx", "w", zipfile.ZIP_STORED) as zout:
    for name, data in payload.items():
        zout.writestr(name, data)

partial_payload = dict(payload)
partial_payload["xl/worksheets/sheet1.xml"] = partial_payload[
    "xl/worksheets/sheet1.xml"
].replace(EXT_LIST_BOTH, EXT_LIST_KEEP)
with zipfile.ZipFile("partial-extension.xlsx", "w", zipfile.ZIP_STORED) as zout:
    for name, data in partial_payload.items():
        zout.writestr(name, data)

EXT_TAG = f"{{{SHEET_NS}}}ext"


def normalized_text(value):
    return (value or "").strip()


def normalized_element(element):
    return (
        element.tag,
        tuple(sorted(element.attrib.items())),
        normalized_text(element.text),
        tuple((normalized_element(child), normalized_text(child.tail)) for child in element),
    )


def worksheet_extension_records(archive, info):
    records = Counter()
    extension_depth = 0
    with archive.open(info) as stream:
        for event, element in ET.iterparse(stream, events=("start", "end")):
            if event == "start":
                if extension_depth:
                    extension_depth += 1
                elif element.tag == EXT_TAG:
                    extension_depth = 1
            elif extension_depth:
                if extension_depth == 1:
                    children = tuple(
                        (normalized_element(child), normalized_text(child.tail))
                        for child in element
                    )
                    records[(element.attrib.get("uri", ""), children)] += 1
                    element.clear()
                    extension_depth = 0
                else:
                    extension_depth -= 1
            else:
                element.clear()
    return records


def archive_inventory(source):
    with zipfile.ZipFile(source) as archive:
        names = set(archive.namelist())
        extensions = {}
        for info in archive.infolist():
            if (info.filename.startswith("xl/worksheets/")
                    and info.filename.endswith(".xml")):
                extensions[info.filename] = worksheet_extension_records(archive, info)
        return names, extensions


def stripped_extension_records(before, after, common_names):
    stripped = []
    for name in sorted(common_names):
        for (uri, children), count in (before.get(name, Counter())
                                       - after.get(name, Counter())).items():
            stripped.extend((name, uri, children) for _ in range(count))
    return sorted(stripped, key=repr)


def load_with_round_trip_audit_from_source(source, **load_options):
    require(not load_options.get("read_only"),
            "round-trip audit requires a normal writable Workbook")
    require(load_options.get("rich_text", True) is True,
            "round-trip audit must preserve rich-text cell runs")
    load_options["rich_text"] = True
    source.seek(0)
    before_names, before_extensions = archive_inventory(source)
    source.seek(0)
    audit_workbook = openpyxl.load_workbook(source, **load_options)
    try:
        with TemporaryFile() as output:
            audit_workbook.save(output)
            output.seek(0)
            after_names, after_extensions = archive_inventory(output)
    finally:
        audit_workbook.close()
    dropped = sorted(before_names - after_names)
    stripped_extensions = stripped_extension_records(
        before_extensions, after_extensions, before_names & after_names
    )
    source.seek(0)
    editable_workbook = openpyxl.load_workbook(source, **load_options)
    return editable_workbook, dropped, stripped_extensions


def load_with_round_trip_audit(path, **load_options):
    with validated_xlsx_source(path) as source:
        return load_with_round_trip_audit_from_source(source, **load_options)


audited_extended, dropped, stripped = load_with_round_trip_audit("extended.xlsx")
audited_extended.close()
check("injected slicer-like part is detected as dropped", "xl/slicers/slicer1.xml" in dropped, dropped)
check("each stripped worksheet extension record is reported with worksheet and URI",
      sum(item[0] == "xl/worksheets/sheet1.xml" and item[1] == EXT_URI
          for item in stripped) == 2,
      stripped)
audited_plain, clean_dropped, clean_stripped = load_with_round_trip_audit("plain.xlsx")
check("clean workbook reports nothing", (clean_dropped, clean_stripped) == ([], []))
check("round-trip audit returns the actual editable workbook identity",
      audited_plain["Data"]["A2"].value == "EU")
audited_plain.close()

rich_input = openpyxl.Workbook()
rich_input.active.title = "Rich"
rich_input.active["A1"] = CellRichText(
    TextBlock(InlineFont(b=True), "Bold"), " and plain"
)
rich_input.save("rich-roundtrip.xlsx")
rich_editable, rich_dropped, rich_stripped = load_with_round_trip_audit(
    "rich-roundtrip.xlsx"
)
rich_editable["Rich"]["B1"] = "edited"
rich_editable.save("rich-roundtrip-edited.xlsx")
rich_editable.close()
rich_reopened = openpyxl.load_workbook("rich-roundtrip-edited.xlsx", rich_text=True)
rich_value = rich_reopened["Rich"]["A1"].value
check(
    "round-trip audit preserves rich-text runs while editing another cell",
    rich_dropped == [] and rich_stripped == []
    and isinstance(rich_value, CellRichText)
    and isinstance(rich_value[0], TextBlock) and rich_value[0].font.b is True
    and str(rich_value) == "Bold and plain"
    and rich_reopened["Rich"]["B1"].value == "edited",
    (rich_dropped, rich_stripped, type(rich_value), rich_value),
)
rich_reopened.close()
try:
    load_with_round_trip_audit("plain.xlsx", rich_text=False)
    rich_text_opt_out_rejected = False
except ValueError as error:
    rich_text_opt_out_rejected = "must preserve rich-text" in str(error)
check("round-trip audit rejects rich-text flattening opt-outs", rich_text_opt_out_rejected)

from openpyxl.drawing.image import Image as AuditImage
Path("audit-image.png").write_bytes(base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAusB9Y9ZlRYAAAAASUVORK5CYII="
))
image_input_wb = openpyxl.Workbook()
image_input_wb.active.title = "Data"
image_input_wb.active["A1"] = "before"
image_input_wb.active.add_image(AuditImage("audit-image.png"), "C3")
image_input_wb.save("image-roundtrip.xlsx")
image_editable, image_dropped, image_stripped = load_with_round_trip_audit(
    "image-roundtrip.xlsx"
)
image_editable["Data"]["A1"] = "after"
image_editable.save("image-roundtrip-edited.xlsx")
image_editable.close()
image_reopened = openpyxl.load_workbook("image-roundtrip-edited.xlsx")
check("round-trip audit returns a fresh workbook whose image streams can be saved again",
      image_dropped == [] and image_stripped == []
      and image_reopened["Data"]["A1"].value == "after"
      and len(image_reopened["Data"]._images) == 1,
      (image_dropped, image_stripped, len(image_reopened["Data"]._images)))
image_reopened.close()

real_archive_inventory = archive_inventory
preflight_inventory_calls = []
def tracking_archive_inventory(source):
    preflight_inventory_calls.append(source)
    return real_archive_inventory(source)
archive_inventory = tracking_archive_inventory
try:
    load_with_round_trip_audit("compressed-xlsx-bomb.xlsx")
    audit_bomb_rejected = False
except ValueError as error:
    audit_bomb_rejected = "suspicious compression ratio: xl/styles.xml" in str(error)
finally:
    archive_inventory = real_archive_inventory
check("round-trip audit rejects a package bomb before raw inventory",
      audit_bomb_rejected and preflight_inventory_calls == [],
      (audit_bomb_rejected, preflight_inventory_calls))
inventory_names, inventory_extensions = archive_inventory("extended.xlsx")
check("binary archive parts are named but never marker-scanned",
      "xl/media/large.bin" in inventory_names
      and "xl/media/large.bin" not in inventory_extensions)

partial_names, partial_extensions = archive_inventory("partial-extension.xlsx")
coarse_needles = (b"extLst", X14_NS.encode(), EXT_URI.encode())
coarse_before = {needle for needle in coarse_needles
                 if needle in payload["xl/worksheets/sheet1.xml"]}
coarse_after = {needle for needle in coarse_needles
                if needle in partial_payload["xl/worksheets/sheet1.xml"]}
check("coarse marker presence misses one removed extension record (negative control)",
      coarse_before == coarse_after == set(coarse_needles),
      (coarse_before, coarse_after))
partial_loss = stripped_extension_records(
    inventory_extensions, partial_extensions, inventory_names & partial_names
)
check("granular extension inventory detects only the removed child-content record",
      len(partial_loss) == 1
      and partial_loss[0][0] == "xl/worksheets/sheet1.xml"
      and partial_loss[0][1] == EXT_URI
      and "drop" in repr(partial_loss[0][2]),
      partial_loss)
check("extension inventory keeps records separated by worksheet identity",
      inventory_extensions["xl/worksheets/sheet2.xml"]
      == partial_extensions["xl/worksheets/sheet2.xml"]
      and bool(partial_extensions["xl/worksheets/sheet2.xml"]),
      partial_extensions)

# ---- formatting.md snippet: sheet references built from the real sheet title -------
wb_f = openpyxl.Workbook()
src = wb_f.active
src.title = "Raw Data"          # space forces quoting
src.append(["Region", "Units", "Price"])
src.append(["EU", 3, 10])
src.append(["US", 4, 20])
agg = wb_f.create_sheet("ByRegion")


def sheet_ref(sheet):
    escaped = sheet.title.replace("'", "''")
    return f"'{escaped}'!"


ref = sheet_ref(src)
agg.append(["Region", "Units"])
agg["A2"] = "EU"
agg["B2"] = f"=SUMIF({ref}A:A,A2,{ref}B:B)"
wb_f.save("agg.xlsx")
wb_g = openpyxl.load_workbook("agg.xlsx")
check("formula references the real sheet name", wb_g["ByRegion"]["B2"].value == "=SUMIF('Raw Data'!A:A,A2,'Raw Data'!B:B)", wb_g["ByRegion"]["B2"].value)
apostrophe_sheet = wb_f.create_sheet("O'Brien")
apostrophe_sheet["A1"] = 1
agg["B3"] = f"=SUM({sheet_ref(apostrophe_sheet)}A:A)"
wb_f.save("apostrophe-agg.xlsx")
apostrophe_formula = openpyxl.load_workbook("apostrophe-agg.xlsx")["ByRegion"]["B3"].value
check(
    "quoted sheet reference doubles apostrophes",
    apostrophe_formula == "=SUM('O''Brien'!A:A)",
    apostrophe_formula,
)
hyphen_sheet = wb_f.create_sheet("Q1-Data")
hyphen_sheet["A1"] = 1
agg["B4"] = f"=SUM({sheet_ref(hyphen_sheet)}A:A)"
wb_f.save("hyphen-agg.xlsx")
hyphen_formula = openpyxl.load_workbook("hyphen-agg.xlsx")["ByRegion"]["B4"].value
check("ambiguous punctuation is protected by quoting", hyphen_formula == "=SUM('Q1-Data'!A:A)", hyphen_formula)

# Falsey values are valid categories; blank filtering must not discard or conflate them.
falsey_ws = openpyxl.Workbook().active
for value in ("Category", 0, False, "", None, 0, False):
    falsey_ws.append([value])
regions = []
seen_region_keys = set()
for (region,) in falsey_ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
    if region is None or region == "":
        continue
    key = (type(region), region)
    if key not in seen_region_keys:
        seen_region_keys.add(key)
        regions.append(region)
check(
    "aggregation preserves numeric zero and boolean false as distinct categories",
    len(regions) == 2
    and type(regions[0]) is int and regions[0] == 0
    and type(regions[1]) is bool and regions[1] is False,
    [(type(value).__name__, value) for value in regions],
)

# ---- edit.md structural audit includes non-cell dependencies ------------------
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.formula import Tokenizer
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.table import Table


# ---- create.md chart: duplicate products are described as row-level sales ----
row_chart_wb = openpyxl.Workbook()
row_chart_ws = row_chart_wb.active
row_chart_ws.title = "Sales"
row_chart_ws.append(["Region", "Product", "Revenue"])
for row in (("EU", "Widget", 1140), ("EU", "Gadget", 1680), ("US", "Widget", 1900)):
    row_chart_ws.append(row)
row_chart = BarChart()
row_chart.type = "col"
row_chart.title = "Revenue by transaction row"
row_chart.add_data(Reference(row_chart_ws, min_col=3, min_row=1, max_row=4),
                   titles_from_data=True)
row_chart.set_categories(Reference(row_chart_ws, min_col=2, min_row=2, max_row=4))
row_chart_ws.add_chart(row_chart, "E2")
row_chart_wb.save("row-level-sales-chart.xlsx")
row_chart_reopened = openpyxl.load_workbook("row-level-sales-chart.xlsx", data_only=False)
reopened_row_chart = row_chart_reopened["Sales"]._charts[0]
row_chart_title = reopened_row_chart.title.tx.rich.p[0].r[0].t
check("duplicate product labels are explicitly charted per transaction row",
      [row_chart_reopened["Sales"][f"B{row}"].value for row in range(2, 5)].count("Widget") == 2
      and row_chart_title == "Revenue by transaction row"
      and reopened_row_chart.ser[0].cat.numRef.f == "'Sales'!$B$2:$B$4"
      and reopened_row_chart.ser[0].val.numRef.f == "'Sales'!$C$2:$C$4"
      and reopened_row_chart.ser[0].tx.strRef.f == "'Sales'!C1",
      row_chart_title)
row_chart_reopened.close()


def formula_text(value):
    if isinstance(value, str):
        return value
    if text := getattr(value, "text", None):
        return text
    fields = ("ref", "r1", "r2", "dt2D", "dtr", "ca", "del1", "del2")
    details = ", ".join(
        f"{name}={getattr(value, name)!r}" for name in fields if hasattr(value, name)
    )
    return f"{type(value).__name__}({details})"


def defined_name_values(workbook):
    names = workbook.defined_names
    return names.values() if hasattr(names, "values") else names.definedName


def drawing_anchor_rows(drawing):
    anchor = drawing.anchor
    if isinstance(anchor, str):
        return (coordinate_to_tuple(anchor)[0],)
    rows = []
    if marker := getattr(anchor, "_from", None):
        rows.append(marker.row + 1)
    if marker := getattr(anchor, "to", None):
        rows.append(marker.row + 1)
    return tuple(rows)


def sparse_cells(sheet):
    if not hasattr(sheet, "_cells"):
        raise RuntimeError("structural edits require a normal writable Worksheet")
    return sorted(sheet._cells.values(), key=lambda cell: (cell.row, cell.column))


def sparse_formula_cells(sheet):
    return (cell for cell in sparse_cells(sheet) if cell.data_type == "f")


def structural_references(workbook):
    refs = []
    for item in defined_name_values(workbook):
        refs.append(("defined name", item.name, item.attr_text))
    for sheet in workbook.worksheets:
        owner = sheet.title
        for cell in sparse_formula_cells(sheet):
            refs.append(("cell formula", f"{owner}!{cell.coordinate}",
                         formula_text(cell.value)))
        for cell in sparse_cells(sheet):
            hyperlink = cell.hyperlink
            if hyperlink is None:
                continue
            location = getattr(hyperlink, "location", None)
            target = getattr(hyperlink, "target", None)
            refs.append((
                "cell hyperlink",
                f"{owner}!{cell.coordinate}",
                (getattr(hyperlink, "ref", None), location, target),
            ))
        for table in sheet.tables.values():
            refs.append(("table", owner + "!" + table.name, table.ref))
        for merged_range in sheet.merged_cells.ranges:
            refs.append(("merged range", owner, str(merged_range)))
        if sheet.auto_filter.ref:
            refs.append(("auto filter", owner, sheet.auto_filter.ref))
        for label, value in (
            ("print area", sheet.print_area),
            ("print title rows", sheet.print_title_rows),
            ("print title columns", sheet.print_title_cols),
        ):
            if value:
                refs.append((label, owner, str(value)))
        for validation in sheet.data_validations.dataValidation:
            refs.append(("data validation range", owner, str(validation.sqref)))
            for formula in (validation.formula1, validation.formula2):
                if formula:
                    refs.append(("data validation formula", owner, str(formula)))
        for conditional_range in sheet.conditional_formatting:
            refs.append(("conditional formatting range", owner, str(conditional_range.sqref)))
            for rule in sheet.conditional_formatting[conditional_range]:
                for formula in getattr(rule, "formula", ()):
                    refs.append(("conditional formatting formula", owner, str(formula)))
        for index, chart in enumerate(sheet._charts, start=1):
            refs.append(("drawing anchor", f"{owner} chart {index}", drawing_anchor_rows(chart)))
            for element in chart._write().iter():
                if element.tag.rsplit("}", 1)[-1] == "f" and element.text:
                    refs.append(("chart series", f"{owner} chart {index}", element.text))
        for index, image in enumerate(sheet._images, start=1):
            refs.append(("drawing anchor", f"{owner} image {index}", drawing_anchor_rows(image)))
    return refs


def non_cell_references(workbook):
    return [reference for reference in structural_references(workbook)
            if reference[0] != "cell formula"]


def cell_formula_references(workbook):
    refs = []
    for sheet in workbook.worksheets:
        for cell in sparse_formula_cells(sheet):
            value = cell.value
            refs.append((
                "cell formula",
                sheet.title,
                cell.coordinate,
                formula_text(value),
            ))
    return refs


def formula_may_intersect_rows(owner_sheet, formula, shifted_sheet, start_row):
    if not isinstance(formula, str) or not formula.startswith("="):
        return True
    tokens = Tokenizer(formula).items
    unmodeled_reference_functions = {"indirect", "offset", "address", "hyperlink"}
    if any(
        token.type == "FUNC" and token.subtype == "OPEN"
        and token.value.rstrip("(").rsplit(":", 1)[-1]
        .lstrip("@").rsplit(".", 1)[-1].casefold()
        in unmodeled_reference_functions
        for token in tokens
    ):
        return True
    for token in tokens:
        if token.type != "OPERAND" or token.subtype != "RANGE":
            continue
        reference = token.value
        target_sheet = owner_sheet
        if "!" in reference:
            qualifier, reference = reference.rsplit("!", 1)
            if "[" in qualifier or ":" in qualifier:
                return True
            target_sheet = qualifier.strip("'").replace("''", "'")
        if target_sheet.casefold() != shifted_sheet.casefold():
            continue
        try:
            _, min_row, _, max_row = range_boundaries(reference.replace("$", ""))
        except ValueError:
            return True
        if min_row is None or max_row is None or max_row >= start_row:
            return True
    return False

class LegacyDefinedNames:
    """Minimal openpyxl 3.0-style DefinedNameList surface."""
    definedName = [DefinedName("LegacyRange", attr_text="'Legacy'!$A$1")]


class LegacyWorkbook:
    defined_names = LegacyDefinedNames()
    worksheets = []


check(
    "structural audit supports openpyxl 3.0 DefinedNameList",
    structural_references(LegacyWorkbook())
    == [("defined name", "LegacyRange", "'Legacy'!$A$1")],
)


audit_wb = openpyxl.Workbook()
audit_ws = audit_wb.active
audit_ws.title = "Audit"
audit_ws.append(["Value"])
audit_ws.append([1])
audit_ws.append([2])
audit_ws["C1"] = "=SUM(A2:A3)"
audit_wb.defined_names.add(DefinedName("AuditRange", attr_text="'Audit'!$A$2:$A$3"))
audit_ws.add_table(Table(displayName="AuditTable", ref="A1:A3"))
audit_ws.merge_cells("B2:B3")
audit_ws.auto_filter.ref = "A1:A3"
audit_ws.print_area = "A1:B3"
audit_ws.print_title_rows = "1:1"
audit_ws.print_title_cols = "A:A"
validation = DataValidation(type="whole", formula1="'Audit'!$A$2")
validation.add("A2:A3")
audit_ws.add_data_validation(validation)
audit_ws.conditional_formatting.add("A2:A3", FormulaRule(formula=["A2>0"]))
chart = BarChart()
chart.add_data(Reference(audit_ws, min_col=1, min_row=1, max_row=3), titles_from_data=True)
audit_ws.add_chart(chart, "C1")
with open("anchor.png", "wb") as stream:
    stream.write(base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAusB9Y9ZlRYAAAAASUVORK5CYII="
    ))
audit_ws.add_image(Image("anchor.png"), "E5")
audit_references = structural_references(audit_wb)
reference_kinds = {kind for kind, _, _ in audit_references}
formula_references = cell_formula_references(audit_wb)
check(
    "structural audit covers names, tables, filters, validation, formatting, and charts",
    {"defined name", "cell formula", "table", "merged range", "auto filter", "print area",
     "print title rows", "print title columns", "data validation range",
     "data validation formula", "conditional formatting range",
     "conditional formatting formula", "chart series", "drawing anchor"} <= reference_kinds,
    reference_kinds,
)
check("drawing audit records an image anchored at the insertion row",
      ("drawing anchor", "Audit image 1", (5,)) in audit_references, audit_references)
check(
    "structural audit snapshots ordinary cell formulas before row insertion",
    formula_references == [("cell formula", "Audit", "C1", "=SUM(A2:A3)")],
    formula_references,
)

hyperlink_wb = openpyxl.Workbook()
hyperlink_ws = hyperlink_wb.active
hyperlink_ws.title = "Data"
hyperlink_ws["A10"] = "destination"
hyperlink_ws["B2"] = "location link"
hyperlink_ws["B2"].hyperlink = Hyperlink(ref="B2", location="'Data'!A10")
hyperlink_ws["B3"] = "hash target link"
hyperlink_ws["B3"].hyperlink = Hyperlink(ref="B3", target="#'Data'!A10")
hyperlink_ws["B4"] = "external link"
hyperlink_ws["B4"].hyperlink = "https://example.com/"
hyperlink_references = non_cell_references(hyperlink_wb)
check(
    "structural inventory records every hyperlink anchor and destination form",
    ("cell hyperlink", "Data!B2", ("B2", "'Data'!A10", None))
    in hyperlink_references
    and ("cell hyperlink", "Data!B3", ("B3", None, "#'Data'!A10"))
    in hyperlink_references
    and ("cell hyperlink", "Data!B4", ("B4", None, "https://example.com/"))
    in hyperlink_references,
    hyperlink_references,
)
check("cell hyperlinks block structural edits before insert_rows runs",
      bool(hyperlink_references) and hyperlink_ws["A10"].value == "destination")

# Negative controls prove why both destination and anchor metadata need a rewrite plan.
destination_stale_wb = openpyxl.Workbook()
destination_stale_ws = destination_stale_wb.active
destination_stale_ws.title = "Data"
destination_stale_ws["A10"] = "destination"
destination_stale_ws["B2"] = "jump"
destination_stale_ws["B2"].hyperlink = Hyperlink(ref="B2", location="'Data'!A10")
destination_stale_ws.insert_rows(5)
check("insert_rows leaves an internal hyperlink destination stale (negative control)",
      destination_stale_ws["A11"].value == "destination"
      and destination_stale_ws["B2"].hyperlink.location == "'Data'!A10")

anchor_stale_wb = openpyxl.Workbook()
anchor_stale_ws = anchor_stale_wb.active
anchor_stale_ws.title = "Data"
anchor_stale_ws["B8"] = "moving link"
anchor_stale_ws["B8"].hyperlink = Hyperlink(ref="B8", target="https://example.com/")
anchor_stale_ws.insert_rows(5)
check("insert_rows leaves an external hyperlink anchor ref stale (negative control)",
      anchor_stale_ws["B9"].value == "moving link"
      and anchor_stale_ws["B9"].hyperlink.ref == "B8")
anchor_stale_wb.save("stale-hyperlink-anchor.xlsx")
anchor_stale_reopened = openpyxl.load_workbook("stale-hyperlink-anchor.xlsx")
check("stale hyperlink ref detaches from moved text after save/reopen (negative control)",
      anchor_stale_reopened["Data"]["B8"].hyperlink is not None
      and anchor_stale_reopened["Data"]["B9"].value == "moving link"
      and anchor_stale_reopened["Data"]["B9"].hyperlink is None)
anchor_stale_reopened.close()

sparse_scan_wb = openpyxl.Workbook()
sparse_scan_ws = sparse_scan_wb.active
sparse_scan_ws.title = "Sparse"
sparse_scan_ws["D2"] = "=1+1"
sparse_scan_ws["XFD1048576"].number_format = "0.00"
sparse_scan_wb.save("sparse-structural-scan.xlsx")
sparse_scan_wb.close()
sparse_scan_wb = openpyxl.load_workbook("sparse-structural-scan.xlsx", data_only=False)
sparse_scan_ws = sparse_scan_wb["Sparse"]
original_sparse_iter_rows = sparse_scan_ws.iter_rows
sparse_scan_ws.iter_rows = lambda *args, **kwargs: (_ for _ in ()).throw(
    RuntimeError("rectangular formula scan must not run")
)
try:
    sparse_structural = structural_references(sparse_scan_wb)
    sparse_formulas = cell_formula_references(sparse_scan_wb)
finally:
    sparse_scan_ws.iter_rows = original_sparse_iter_rows
check("structural formula inventory walks sparse cells at worksheet limits",
      ("cell formula", "Sparse!D2", "=1+1") in sparse_structural
      and sparse_formulas == [("cell formula", "Sparse", "D2", "=1+1")],
      (sparse_structural, sparse_formulas))
sparse_scan_wb.close()
check("intersecting formula ranges are blocked before row insertion",
      formula_may_intersect_rows("Audit", "=SUM(A2:A3)", "Audit", 3))
check("audited formula ranges above the insertion can proceed",
      not formula_may_intersect_rows("Data", "=C2*1.08", "Data", 5))
check("sheet qualifiers are matched case-insensitively",
      formula_may_intersect_rows("Summary", "=SUM(data!A5:A6)", "Data", 5))
check("a genuinely different sheet remains outside the shifted rows",
      not formula_may_intersect_rows("Summary", "=SUM(Archive!A5:A6)", "Data", 5))
check("INDIRECT string references require a manual structural rewrite plan",
      formula_may_intersect_rows("Data", '=SUM(INDIRECT("A5:A6"))', "Data", 5))
check("OFFSET numeric row references require a manual structural rewrite plan",
      formula_may_intersect_rows("Data", "=SUM(OFFSET(A1,4,0,2,1))", "Data", 5))
check("HYPERLINK string destinations require a manual structural rewrite plan",
      formula_may_intersect_rows(
          "Data", '=HYPERLINK("#\'Data\'!A10","jump")', "Data", 5
      ))
check("implicit-intersection INDIRECT references require a manual rewrite plan",
      formula_may_intersect_rows("Data", '=@INDIRECT("A5:A6")', "Data", 5))
check("OFFSET used by the range operator requires a manual rewrite plan",
      formula_may_intersect_rows("Data", "=SUM(A1:OFFSET(A1,5,0))", "Data", 5))
check("INDIRECT text inside a string does not create a dynamic reference",
      not formula_may_intersect_rows(
          "Data", '=IF(A1="INDIRECT(A5:A6)",1,0)', "Data", 5
      ))

stale_wb = openpyxl.Workbook()
stale_ws = stale_wb.active
stale_ws["A5"], stale_ws["A6"] = 10, 20
stale_ws["B1"] = "=SUM(A5:A6)"
stale_formula_before = cell_formula_references(stale_wb)
stale_ws.insert_rows(5)
check(
    "insert_rows leaves intersecting formulas stale (negative control)",
    stale_formula_before == [("cell formula", "Sheet", "B1", "=SUM(A5:A6)")]
    and stale_ws["B1"].value == "=SUM(A5:A6)"
    and (stale_ws["A6"].value, stale_ws["A7"].value) == (10, 20),
    (stale_formula_before, stale_ws["B1"].value),
)

safe_wb = openpyxl.Workbook()
safe_ws = safe_wb.active
safe_ws.title = "Data"
safe_ws["C2"], safe_ws["D2"] = 100, "=C2*1.08"
safe_before = cell_formula_references(safe_wb)
safe_dependencies = [
    reference for reference in safe_before
    if formula_may_intersect_rows(reference[1], reference[3], "Data", 5)
]
if not safe_dependencies:
    safe_ws.insert_rows(5)
    safe_wb.calculation.fullCalcOnLoad = True
    safe_wb.calculation.forceFullCalc = True
    safe_wb.calculation.calcMode = "auto"
    safe_wb.save("audited-structural-edit.xlsx")
safe_reopened = openpyxl.load_workbook("audited-structural-edit.xlsx", data_only=False)
check("audited non-intersecting formula path reaches save",
      safe_reopened["Data"]["D2"].value == "=C2*1.08")
check("formula edit forces recalculation after save",
      safe_reopened.calculation.fullCalcOnLoad is True
      and safe_reopened.calculation.calcMode == "auto")

legacy_name = DefinedName("LegacyName", attr_text="Audit!$A$1")
legacy_names = type("LegacyDefinedNames", (), {"definedName": [legacy_name]})()
legacy_workbook = type("LegacyWorkbook", (), {"defined_names": legacy_names})()
check("defined-name adapter supports openpyxl 3.0 collections",
      list(defined_name_values(legacy_workbook)) == [legacy_name])

formula_only = openpyxl.Workbook()
formula_only.active["A1"] = "=Data!A5"
check("structural guard sees formulas before row insertion",
      structural_references(formula_only)
      == [("cell formula", "Sheet!A1", "=Data!A5")],
      structural_references(formula_only))

audited_wb = openpyxl.Workbook()
audited_ws = audited_wb.active
audited_ws.title = "Data"
audited_ws["B2"] = "=A2*2"
audited_ws["A5"] = "shift me"
audited_references_before = structural_references(audited_wb)
audited_unchanged = {("cell formula", "Data!B2", "=A2*2")}
unaudited_references = [
    reference for reference in audited_references_before
    if reference not in audited_unchanged
]
if not unaudited_references:
    audited_ws.insert_rows(5)
    audited_ws["D2"] = "=C2*1.08"
    audited_wb.save("audited-edit.xlsx")
audited_reopened = openpyxl.load_workbook("audited-edit.xlsx", data_only=False)
check("exact dependency allowlist lets an audited edit reach save",
      audited_references_before == [("cell formula", "Data!B2", "=A2*2")]
      and unaudited_references == []
      and audited_reopened["Data"]["A6"].value == "shift me"
      and audited_reopened["Data"]["B2"].value == "=A2*2"
      and audited_reopened["Data"]["D2"].value == "=C2*1.08")

unsafe_wb = openpyxl.Workbook()
unsafe_ws = unsafe_wb.active
unsafe_ws.title = "Data"
unsafe_ws["A5"], unsafe_ws["A6"] = 10, 20
unsafe_ws["B1"] = "=SUM(A5:A6)"
unsafe_dependencies = [
    reference for reference in structural_references(unsafe_wb)
    if reference not in audited_unchanged
]
unsafe_blocked_before_insert = bool(unsafe_dependencies)
check("intersecting formula not on the exact allowlist is blocked before insertion",
      unsafe_blocked_before_insert
      and unsafe_ws["A5"].value == 10 and unsafe_ws["A6"].value == 20,
      unsafe_dependencies)

# and the edit itself still works after the warning path
wb2 = openpyxl.load_workbook("plain.xlsx")
wb2["Data"]["B2"] = "=B2*1"  # formula stays a formula
wb2["Data"]["B2"].number_format = "#,##0.00"
wb2.save("edited.xlsx")
wb3 = openpyxl.load_workbook("edited.xlsx")
check("edited cell keeps a formula string", isinstance(wb3["Data"]["B2"].value, str) and wb3["Data"]["B2"].value.startswith("="))
expected_number_formats = {"Data": {"B2": "#,##0.00"}}
format_matches = all(
    wb3[sheet][coordinate].number_format == expected
    for sheet, cells in expected_number_formats.items()
    for coordinate, expected in cells.items()
)
check("task-specific number format mapping is verified", format_matches)

# ---- read.md snippet: multi-sheet profiles cover every sheet ----------------------
import posixpath
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CELL_TAG = f"{{{MAIN_NS}}}c"
FORMULA_TAG = f"{{{MAIN_NS}}}f"
VALUE_TAG = f"{{{MAIN_NS}}}v"
INLINE_STRING_TAG = f"{{{MAIN_NS}}}is"
MAX_EXPLICIT_CELLS = 1_000_000
MAX_PROFILE_RECTANGLE_CELLS = 100_000


def worksheet_part(archive, sheet_name):
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    sheet = next(
        item for item in workbook.iter(f"{{{MAIN_NS}}}sheet")
        if item.attrib["name"] == sheet_name
    )
    relationship_id = sheet.attrib[f"{{{DOC_REL_NS}}}id"]
    relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    target = next(
        item.attrib["Target"] for item in relationships.iter(f"{{{PKG_REL_NS}}}Relationship")
        if item.attrib["Id"] == relationship_id
    )
    part = target.lstrip("/") if target.startswith("/") else posixpath.normpath(
        posixpath.join("xl", target)
    )
    require(part in archive.namelist(), f"worksheet part is missing: {part}")
    return part


def worksheet_xml_profile(archive, part):
    min_row = min_column = max_row = max_column = None
    explicit_cell_count = 0
    formula_count = missing_formula_count = 0
    missing_formula_samples = []
    coordinate = cell_type = value_text = formula_display = None
    has_formula = value_seen = inline_string_seen = False
    with archive.open(part) as source:
        for event, element in ET.iterparse(source, events=("start", "end")):
            if event == "start" and element.tag == CELL_TAG:
                coordinate = element.attrib.get("r")
                require(coordinate is not None, f"cell without a coordinate in {part}")
                cell_type = element.attrib.get("t")
                has_formula = value_seen = inline_string_seen = False
                value_text = formula_display = None
            elif event == "end" and coordinate is not None:
                if element.tag == FORMULA_TAG:
                    has_formula = True
                    if element.text is not None:
                        formula_display = "=" + element.text
                    else:
                        details = ", ".join(
                            f"{key}={value!r}" for key, value in sorted(element.attrib.items())
                        )
                        formula_display = f"<f {details}>"
                elif element.tag == VALUE_TAG:
                    value_seen = True
                    value_text = element.text
                elif element.tag == INLINE_STRING_TAG and cell_type == "inlineStr":
                    inline_string_seen = True
                elif element.tag == CELL_TAG:
                    explicit_cell_count += 1
                    require(explicit_cell_count <= MAX_EXPLICIT_CELLS,
                            f"too many explicit worksheet cells in {part}")
                    row_index, column_index = coordinate_to_tuple(coordinate)
                    require(1 <= row_index <= 1_048_576 and 1 <= column_index <= 16_384,
                            f"cell coordinate outside XLSX limits: {coordinate}")
                    scalar_value = value_seen and (
                        value_text not in (None, "") or cell_type == "str"
                    )
                    populated = has_formula or scalar_value or inline_string_seen
                    if populated:
                        min_row = row_index if min_row is None else min(min_row, row_index)
                        min_column = (column_index if min_column is None
                                      else min(min_column, column_index))
                        max_row = row_index if max_row is None else max(max_row, row_index)
                        max_column = (column_index if max_column is None
                                      else max(max_column, column_index))
                    valid_cache = value_seen and (
                        value_text not in (None, "") or cell_type == "str"
                    )
                    if has_formula:
                        formula_count += 1
                        if not valid_cache:
                            missing_formula_count += 1
                            if len(missing_formula_samples) < 10:
                                missing_formula_samples.append((coordinate, formula_display))
                    coordinate = None
                element.clear()
            elif event == "end":
                element.clear()
    if max_row is None:
        bounds, extent, first_populated_row = None, "A1:A1", None
    else:
        bounds = (min_row, min_column, max_row, max_column)
        extent = (f"{get_column_letter(min_column)}{min_row}:"
                  f"{get_column_letter(max_column)}{max_row}")
        first_populated_row = min_row
    return {
        "part": part,
        "bounds": bounds,
        "extent": extent,
        "first_populated_row": first_populated_row,
        "formula_count": formula_count,
        "missing_formula_count": missing_formula_count,
        "missing_formula_samples": missing_formula_samples,
        "explicit_cell_count": explicit_cell_count,
    }


def bounded_sample_rows(worksheet, profile, *, max_cells=MAX_PROFILE_RECTANGLE_CELLS):
    if profile["bounds"] is None:
        return iter(())
    min_row, min_column, max_row, max_column = profile["bounds"]
    sample_max_row = min(max_row, min_row + 5)
    sample_cells = (sample_max_row - min_row + 1) * (max_column - min_column + 1)
    require(sample_cells <= max_cells,
            f"sample rectangle is too large: {worksheet.title} ({sample_cells} cells)")
    return worksheet.iter_rows(
        min_row=min_row, min_col=min_column,
        max_row=sample_max_row, max_col=max_column,
        values_only=True,
    )


def worksheet_declared_dimension(worksheet):
    if worksheet.max_row is None or worksheet.max_column is None:
        return None
    return worksheet.calculate_dimension()

wb_h = openpyxl.Workbook()
wb_h.active.title = "First"
wb_h.active["A1"] = "=1+1"
wb_h.active["A2"] = ArrayFormula("A2:A3", "=ROW(A2:A3)")
wb_h.active["C1"] = "=literal"
wb_h.active["C1"].data_type = "s"
wb_h.active["D1"] = "+literal"
wb_h.active["E1"] = "-literal"
wb_h.active["F1"] = "@literal"
second = wb_h.create_sheet("Second")
second["A1"] = "plain"
second["A2"] = "=2+2"
second["B1"] = DataTableFormula(ref="B1:B2", r1="C1")
wb_h.save("multi.xlsx")
# openpyxl writes an empty <v/> for uncached formulas. Remove those elements explicitly so
# this fixture represents a truly absent cache rather than a cached displayed blank.
with zipfile.ZipFile("multi.xlsx") as archive:
    multi_members = {name: archive.read(name) for name in archive.namelist()}
for name in [item for item in multi_members if item.startswith("xl/worksheets/")
             and item.endswith(".xml")]:
    root = ET.fromstring(multi_members[name])
    for cell in root.iter(CELL_TAG):
        if cell.find(FORMULA_TAG) is not None:
            cached_value = cell.find(VALUE_TAG)
            if cached_value is not None:
                cell.remove(cached_value)
    multi_members[name] = ET.tostring(root, encoding="utf-8", xml_declaration=True)
with zipfile.ZipFile("multi.xlsx", "w", zipfile.ZIP_DEFLATED) as archive:
    for name, data in multi_members.items():
        archive.writestr(name, data)
with validated_xlsx_source("multi.xlsx") as multi_source:
    formula_wb = openpyxl.load_workbook(multi_source, read_only=True, data_only=False)
    multi_source.seek(0)
    value_wb = openpyxl.load_workbook(multi_source, read_only=True, data_only=True)
    try:
        multi_source.seek(0)
        with zipfile.ZipFile(multi_source) as archive:
            profiled = list(value_wb.sheetnames)
            multi_profiles = {
                sheet_name: worksheet_xml_profile(
                    archive, worksheet_part(archive, sheet_name)
                )
                for sheet_name in profiled
            }
        uncached = {
            (sheet_name, coordinate, formula)
            for sheet_name, profile in multi_profiles.items()
            for coordinate, formula in profile["missing_formula_samples"]
        }
        paired_same_source = (
            next(formula_wb["First"].iter_rows(min_row=1, max_row=1))[0].value == "=1+1"
            and next(value_wb["First"].iter_rows(min_row=1, max_row=1))[0].value is None
        )
    finally:
        formula_wb.close()
        value_wb.close()
check("multi-sheet profile iterates every sheet", profiled == ["First", "Second"], profiled)
check("formula/value read-only streams share one validated source identity",
      paired_same_source)
check("uncached formulas found on both sheets", {item[0] for item in uncached} == {"First", "Second"}, uncached)
check("array-formula objects are detected by data_type", ("First", "A2", "=ROW(A2:A3)") in uncached, uncached)
data_table_entry = next(item for item in uncached if item[:2] == ("Second", "B1"))
check("data-table formulas have stable diagnostic text",
      data_table_entry[2].startswith("<f ")
      and "ref='B1:B2'" in data_table_entry[2]
      and "r1='C1'" in data_table_entry[2], data_table_entry)

# csv.md: value export uses the cached-value workbook and reports every missing cache.
def cached_formula_coordinates(archive, part, wanted_coordinates=None):
    cached = set()
    coordinate = None
    cell_type = None
    has_formula = value_seen = False
    value_text = None
    with archive.open(part) as source:
        for event, element in ET.iterparse(source, events=("start", "end")):
            if event == "start" and element.tag == CELL_TAG:
                coordinate = element.attrib["r"]
                cell_type = element.attrib.get("t")
                has_formula = value_seen = False
                value_text = None
            elif event == "end" and coordinate is not None:
                if element.tag == FORMULA_TAG:
                    has_formula = True
                elif element.tag == VALUE_TAG:
                    value_seen = True
                    value_text = element.text
                elif element.tag == CELL_TAG:
                    valid_cache = value_seen and (
                        value_text not in (None, "") or cell_type == "str"
                    )
                    if (has_formula and valid_cache
                            and (wanted_coordinates is None
                                 or coordinate in wanted_coordinates)):
                        cached.add(coordinate)
                    coordinate = None
                element.clear()
            elif event == "end":
                element.clear()
    return cached


with validated_xlsx_source("multi.xlsx") as multi_csv_source:
    multi_csv_value_wb = openpyxl.load_workbook(
        multi_csv_source, read_only=True, data_only=True
    )
    try:
        multi_csv_source.seek(0)
        with zipfile.ZipFile(multi_csv_source) as archive:
            multi_csv_profile = worksheet_xml_profile(
                archive, worksheet_part(archive, "First")
            )
        missing_caches = [
            coordinate
            for coordinate, _ in multi_csv_profile["missing_formula_samples"]
        ]
    finally:
        multi_csv_value_wb.close()
check("XLSX-to-CSV reports formulas with no cached value before creating output",
      set(missing_caches) >= {"A1", "A2"}, missing_caches)


def write_formula_cache_fixture(path, cache_kind):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = ('=IF(TRUE,"","x")' if cache_kind == "empty-string" else "=1+1")
    workbook.save(path)
    with zipfile.ZipFile(path) as archive:
        members = {name: archive.read(name) for name in archive.namelist()}
    root = ET.fromstring(members["xl/worksheets/sheet1.xml"])
    cell = next(item for item in root.iter(CELL_TAG) if item.attrib["r"] == "A1")
    cached_value = cell.find(VALUE_TAG)
    if cache_kind == "nonempty":
        cell.attrib.pop("t", None)
        if cached_value is None:
            cached_value = ET.SubElement(cell, VALUE_TAG)
        cached_value.text = "2"
    elif cache_kind == "empty-string":
        cell.set("t", "str")
        if cached_value is None:
            cached_value = ET.SubElement(cell, VALUE_TAG)
        cached_value.text = None
    elif cache_kind == "bare-empty":
        cell.attrib.pop("t", None)
        if cached_value is None:
            cached_value = ET.SubElement(cell, VALUE_TAG)
        cached_value.text = None
    elif cache_kind == "absent":
        cell.attrib.pop("t", None)
        if cached_value is not None:
            cell.remove(cached_value)
    else:
        raise ValueError(cache_kind)
    members["xl/worksheets/sheet1.xml"] = ET.tostring(
        root, encoding="utf-8", xml_declaration=True,
    )
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, data in members.items():
            archive.writestr(name, data)


MAX_CSV_EXPORT_CELLS = 5_000_000


def export_formula_values(source_path, output_path, *, max_cells=MAX_CSV_EXPORT_CELLS):
    sheet_name = "Data"
    destination = Path(output_path)
    temporary = None
    try:
        descriptor, temporary_name = mkstemp(
            dir=destination.parent, prefix=f".{destination.name}.", suffix=".tmp"
        )
        temporary = Path(temporary_name)
        os.close(descriptor)
        with validated_xlsx_source(source_path) as package_source:
            value_book = openpyxl.load_workbook(
                package_source, read_only=True, data_only=True
            )
            try:
                package_source.seek(0)
                with zipfile.ZipFile(package_source) as archive:
                    profile = worksheet_xml_profile(
                        archive, worksheet_part(archive, sheet_name)
                    )
                if profile["missing_formula_count"]:
                    raise RuntimeError(
                        "formula cells have no cached value: "
                        f"{profile['missing_formula_samples']}"
                    )
                value_sheet = value_book[sheet_name]
                value_sheet.reset_dimensions()
                if profile["bounds"] is None:
                    rows = iter(())
                else:
                    min_row, min_column, max_row, max_column = profile["bounds"]
                    export_cells = (
                        (max_row - min_row + 1) * (max_column - min_column + 1)
                    )
                    require(export_cells <= max_cells,
                            f"CSV export rectangle is too large: {export_cells} cells")
                    rows = value_sheet.iter_rows(
                        min_row=min_row, min_col=min_column,
                        max_row=max_row, max_col=max_column,
                        values_only=True,
                    )
                with temporary.open("w", newline="", encoding="utf-8") as output:
                    writer = csv.writer(output)
                    for value_row in rows:
                        writer.writerow([
                            spreadsheet_csv_field(value) for value in value_row
                        ])
            finally:
                value_book.close()
    except Exception:
        if temporary is not None:
            temporary.unlink(missing_ok=True)
        raise
    require(temporary is not None, "CSV temporary output was not created")
    temporary.replace(destination)


write_formula_cache_fixture("cached-value.xlsx", "nonempty")
write_formula_cache_fixture("cached-empty.xlsx", "empty-string")
write_formula_cache_fixture("bare-empty-cache.xlsx", "bare-empty")
write_formula_cache_fixture("missing-cache.xlsx", "absent")
with zipfile.ZipFile("cached-value.xlsx") as archive:
    cached_value_cells = cached_formula_coordinates(archive, worksheet_part(archive, "Data"))
with zipfile.ZipFile("cached-empty.xlsx") as archive:
    cached_empty_cells = cached_formula_coordinates(archive, worksheet_part(archive, "Data"))
with zipfile.ZipFile("bare-empty-cache.xlsx") as archive:
    bare_empty_cells = cached_formula_coordinates(archive, worksheet_part(archive, "Data"))
with zipfile.ZipFile("missing-cache.xlsx") as archive:
    truly_missing_cells = cached_formula_coordinates(archive, worksheet_part(archive, "Data"))
check("XML cache inventory accepts nonempty and typed empty-string caches",
      cached_value_cells == cached_empty_cells == {"A1"},
      (cached_value_cells, cached_empty_cells))
check("XML cache inventory rejects bare empty <v/> and an absent cache",
      bare_empty_cells == truly_missing_cells == set(),
      (bare_empty_cells, truly_missing_cells))


def profile_missing_formula_caches(source_path):
    with validated_xlsx_source(source_path) as package_source:
        with zipfile.ZipFile(package_source) as archive:
            profile = worksheet_xml_profile(
                archive, worksheet_part(archive, "Data")
            )
    return [coordinate for coordinate, _ in profile["missing_formula_samples"]]


check("workbook profiling accepts nonempty and typed blank formula caches",
      profile_missing_formula_caches("cached-value.xlsx") == []
      and profile_missing_formula_caches("cached-empty.xlsx") == [])
check("workbook profiling reports absent and untyped empty formula caches",
      profile_missing_formula_caches("missing-cache.xlsx") == ["A1"]
      and profile_missing_formula_caches("bare-empty-cache.xlsx") == ["A1"])


def write_region_key_fixture(path, cache_kind):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Region", "Units", "Revenue", "Source Region"])
    sheet.append(["=D2", 2, 4, "EU"])
    workbook.save(path)
    with zipfile.ZipFile(path) as archive:
        members = {name: archive.read(name) for name in archive.namelist()}
    root = ET.fromstring(members["xl/worksheets/sheet1.xml"])
    cell = next(item for item in root.iter(CELL_TAG) if item.attrib["r"] == "A2")
    value = cell.find(VALUE_TAG)
    if cache_kind == "nonempty":
        cell.set("t", "str")
        if value is None:
            value = ET.SubElement(cell, VALUE_TAG)
        value.text = "EU"
    elif cache_kind == "empty-string":
        cell.set("t", "str")
        if value is None:
            value = ET.SubElement(cell, VALUE_TAG)
        value.text = None
    elif cache_kind == "absent":
        cell.attrib.pop("t", None)
        if value is not None:
            cell.remove(value)
    else:
        raise ValueError(cache_kind)
    members["xl/worksheets/sheet1.xml"] = ET.tostring(
        root, encoding="utf-8", xml_declaration=True,
    )
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, data in members.items():
            archive.writestr(name, data)


def aggregation_regions(source_path, *, max_row_span=100_000,
                        approved_feature_loss=False):
    with validated_xlsx_source(source_path) as package_source:
        formula_book, dropped_parts, stripped_extensions = (
            load_with_round_trip_audit_from_source(package_source, data_only=False)
        )
        if (dropped_parts or stripped_extensions) and not approved_feature_loss:
            formula_book.close()
            raise RuntimeError(
                f"openpyxl would drop parts={dropped_parts!r}, "
                f"extensions={stripped_extensions!r}"
            )
        value_book = None
        try:
            source_sheet = formula_book["Data"]
            source_cells = sorted(
                (cell for cell in source_sheet._cells.values()
                 if cell.row >= 2 and cell.column == 1 and cell.value is not None),
                key=lambda cell: cell.row,
            )
            if source_cells:
                min_source_row, max_source_row = source_cells[0].row, source_cells[-1].row
                row_span = max_source_row - min_source_row + 1
                require(row_span <= max_row_span,
                        f"aggregation row span is too large: {row_span}")
            wanted_formula_coordinates = {
                cell.coordinate for cell in source_cells if cell.data_type == "f"
            }
            package_source.seek(0)
            with zipfile.ZipFile(package_source) as archive:
                cached_cells = cached_formula_coordinates(
                    archive, worksheet_part(archive, "Data"), wanted_formula_coordinates
                )
            package_source.seek(0)
            value_book = openpyxl.load_workbook(
                package_source, read_only=True, data_only=True
            )
            value_sheet = value_book["Data"]
            value_sheet.reset_dimensions()
            regions = []
            missing = []
            if source_cells:
                source_by_row = {cell.row: cell for cell in source_cells}
                value_rows = value_sheet.iter_rows(
                    min_row=min_source_row, max_row=max_source_row,
                    min_col=1, max_col=1,
                )
                for row_index, value_row in enumerate(value_rows, start=min_source_row):
                    source_cell = source_by_row.get(row_index)
                    if source_cell is None:
                        continue
                    value_cell = value_row[0]
                    region = (value_cell.value if source_cell.data_type == "f"
                              else source_cell.value)
                    if (source_cell.data_type == "f" and region is None
                            and source_cell.coordinate not in cached_cells):
                        missing.append(source_cell.coordinate)
                        continue
                    if region is not None and region != "":
                        regions.append(region)
        finally:
            if value_book is not None:
                value_book.close()
            formula_book.close()
    if missing:
        raise RuntimeError(f"aggregation keys have no cached value: {missing}")
    return regions


write_region_key_fixture("region-key-cached.xlsx", "nonempty")
write_region_key_fixture("region-key-empty.xlsx", "empty-string")
write_region_key_fixture("region-key-missing.xlsx", "absent")
cached_regions = aggregation_regions("region-key-cached.xlsx")
aggregate_book = openpyxl.Workbook()
aggregate_book.active["A1"] = cached_regions[0]
check("formula-backed aggregation key is copied from its cached displayed value",
      aggregate_book.active["A1"].value == "EU"
      and aggregate_book.active["A1"].data_type != "f")
check("typed cached blank aggregation keys are skipped",
      aggregation_regions("region-key-empty.xlsx") == [])
try:
    aggregation_regions("region-key-missing.xlsx")
    missing_region_cache_rejected = False
except RuntimeError as error:
    missing_region_cache_rejected = "A2" in str(error)
check("aggregation fails closed when a formula key has no cached value",
      missing_region_cache_rejected)

feature_loss_payload = {
    name: data for name, data in payload.items() if name != "xl/media/large.bin"
}
with zipfile.ZipFile("aggregation-feature-loss.xlsx", "w", zipfile.ZIP_STORED) as archive:
    for name, data in feature_loss_payload.items():
        archive.writestr(name, data)
try:
    aggregation_regions("aggregation-feature-loss.xlsx")
    aggregation_feature_loss_rejected = False
except RuntimeError as error:
    aggregation_feature_loss_rejected = "openpyxl would drop parts=" in str(error)
check("aggregation feature-loss audit gates the exact workbook later edited",
      aggregation_feature_loss_rejected)

export_formula_values("cached-value.xlsx", "cached-value.csv")
with open("cached-value.csv", newline="", encoding="utf-8") as exported:
    cached_value_rows = list(csv.reader(exported))
check("nonempty cached formula result exports its displayed value",
      cached_value_rows == [["2"]], cached_value_rows)
export_formula_values("cached-empty.xlsx", "cached-empty.csv")
with open("cached-empty.csv", newline="", encoding="utf-8") as exported:
    cached_empty_rows = list(csv.reader(exported))
check("cached empty-string formula result exports as a displayed blank",
      cached_empty_rows == [[""]], cached_empty_rows)
Path("missing-cache.csv").write_text("sentinel\n", encoding="utf-8")
Path("missing-cache.csv.tmp").write_text("unrelated temporary\n", encoding="utf-8")
try:
    export_formula_values("missing-cache.xlsx", "missing-cache.csv")
    missing_cache_rejected = False
except RuntimeError as error:
    missing_cache_rejected = "A1" in str(error)
check("formula with no XML cache element is rejected", missing_cache_rejected)
check("failed cache audit preserves the prior destination and removes its temporary file",
      Path("missing-cache.csv").read_text(encoding="utf-8") == "sentinel\n"
      and Path("missing-cache.csv.tmp").read_text(encoding="utf-8")
      == "unrelated temporary\n"
      and list(Path(".").glob(".missing-cache.csv.*.tmp")) == [])
Path("bare-empty-cache.csv").write_text("sentinel\n", encoding="utf-8")
Path("bare-empty-cache.csv.tmp").write_text("unrelated temporary\n", encoding="utf-8")
try:
    export_formula_values("bare-empty-cache.xlsx", "bare-empty-cache.csv")
    bare_empty_rejected = False
except RuntimeError as error:
    bare_empty_rejected = "A1" in str(error)
check("untyped bare empty <v/> from an uncalculated formula is rejected",
      bare_empty_rejected)
check("bare-empty rejection preserves destination and removes temporary output",
      Path("bare-empty-cache.csv").read_text(encoding="utf-8") == "sentinel\n"
      and Path("bare-empty-cache.csv.tmp").read_text(encoding="utf-8")
      == "unrelated temporary\n"
      and list(Path(".").glob(".bare-empty-cache.csv.*.tmp")) == [])

csv_sparse_wb = openpyxl.Workbook()
csv_sparse_ws = csv_sparse_wb.active
csv_sparse_ws.title = "Data"
csv_sparse_ws["A1"], csv_sparse_ws["A2"] = "Region", "EU"
csv_sparse_ws["XFD1048576"].number_format = "0.00"
csv_sparse_wb.save("csv-style-extreme.xlsx")
export_formula_values("csv-style-extreme.xlsx", "csv-style-extreme.csv")
with open("csv-style-extreme.csv", newline="", encoding="utf-8") as exported:
    csv_sparse_rows = list(csv.reader(exported))
check("CSV export ignores a style-only extreme cell without rectangular expansion",
      csv_sparse_rows == [["Region"], ["EU"]], csv_sparse_rows)

csv_far_wb = openpyxl.Workbook()
csv_far_ws = csv_far_wb.active
csv_far_ws.title = "Data"
csv_far_ws["A1"], csv_far_ws["XFD1048576"] = "near", "far"
csv_far_wb.save("csv-far-values.xlsx")
Path("csv-far-values.csv").write_text("sentinel\n", encoding="utf-8")
Path("csv-far-values.csv.tmp").write_text("unrelated temporary\n", encoding="utf-8")
try:
    export_formula_values("csv-far-values.xlsx", "csv-far-values.csv", max_cells=1_000)
    csv_rectangle_rejected = False
except ValueError as error:
    csv_rectangle_rejected = "CSV export rectangle is too large" in str(error)
check("CSV export rejects a far-apart logical rectangle before iteration",
      csv_rectangle_rejected
      and Path("csv-far-values.csv").read_text(encoding="utf-8") == "sentinel\n"
      and Path("csv-far-values.csv.tmp").read_text(encoding="utf-8")
      == "unrelated temporary\n"
      and list(Path(".").glob(".csv-far-values.csv.*.tmp")) == [])

aggregation_sparse_wb = openpyxl.Workbook()
aggregation_sparse_ws = aggregation_sparse_wb.active
aggregation_sparse_ws.title = "Data"
aggregation_sparse_ws["A1"], aggregation_sparse_ws["A2"] = "Region", "EU"
aggregation_sparse_ws["XFD1048576"].number_format = "0.00"
aggregation_sparse_wb.save("aggregation-style-extreme.xlsx")
check("aggregation ignores an extreme style-only cell without iterating a million rows",
      aggregation_regions("aggregation-style-extreme.xlsx") == ["EU"])

aggregation_far_wb = openpyxl.Workbook()
aggregation_far_ws = aggregation_far_wb.active
aggregation_far_ws.title = "Data"
aggregation_far_ws["A1"], aggregation_far_ws["A2"] = "Region", "near"
aggregation_far_ws["A1048576"] = "far"
aggregation_far_wb.save("aggregation-far-values.xlsx")
try:
    aggregation_regions("aggregation-far-values.xlsx", max_row_span=1_000)
    aggregation_span_rejected = False
except ValueError as error:
    aggregation_span_rejected = "aggregation row span is too large" in str(error)
check("aggregation rejects far-apart keys before iter_rows", aggregation_span_rejected)

real_worksheet_xml_profile = worksheet_xml_profile
bomb_profile_calls = []
def tracking_worksheet_xml_profile(*args, **kwargs):
    bomb_profile_calls.append(args)
    return real_worksheet_xml_profile(*args, **kwargs)
worksheet_xml_profile = tracking_worksheet_xml_profile
try:
    export_formula_values("compressed-xlsx-bomb.xlsx", "bomb-export.csv")
    csv_bomb_rejected = False
except ValueError as error:
    csv_bomb_rejected = "suspicious compression ratio: xl/styles.xml" in str(error)
finally:
    worksheet_xml_profile = real_worksheet_xml_profile
check("CSV route rejects a package bomb before raw worksheet inventory",
      csv_bomb_rejected and bomb_profile_calls == [],
      (csv_bomb_rejected, bomb_profile_calls))

# SKILL.md contract: fullCalcOnLoad makes viewers recalculate even in manual calc mode.
calc_wb = openpyxl.Workbook()
calc_ws = calc_wb.active
calc_ws["A1"] = 1
calc_ws["A2"] = 2
calc_ws["A3"] = "=SUM(A1:A2)"
calc_wb.calculation.calcMode = "manual"
calc_wb.calculation.fullCalcOnLoad = False  # simulate a source that does not recalc on load
calc_wb.save("stale-calc.xlsx")
stale_reopened = openpyxl.load_workbook("stale-calc.xlsx")
check("workbook without fullCalcOnLoad round-trips the stale flag (negative control)",
      not bool(getattr(stale_reopened.calculation, "fullCalcOnLoad", False)),
      stale_reopened.calculation)
stale_reopened.calculation.fullCalcOnLoad = True
stale_reopened.save("manual-calc.xlsx")
calc_reopened = openpyxl.load_workbook("manual-calc.xlsx")
check("fullCalcOnLoad survives save/reload",
      bool(getattr(calc_reopened.calculation, "fullCalcOnLoad", False)),
      calc_reopened.calculation)
check("manual calc mode survives save/reload",
      getattr(calc_reopened.calculation, "calcMode", None) == "manual",
      calc_reopened.calculation)
check("reloaded formula cell still holds the formula string",
      calc_reopened.active["A3"].value == "=SUM(A1:A2)", calc_reopened.active["A3"].value)
calc_reopened.close()

# ---- SKILL.md postcheck: formula inventory stays sparse at worksheet limits ----
formula_bound_wb = openpyxl.Workbook()
formula_bound_ws = formula_bound_wb.active
formula_bound_ws["D2"] = "=1+1"
formula_bound_ws["XFD1048576"].number_format = "0.00"  # styled but empty extreme cell
formula_bound_wb.save("formula-bound.xlsx")
formula_bound_reopened = openpyxl.load_workbook("formula-bound.xlsx", data_only=False)
formula_bound_ws = formula_bound_reopened.active


def expected_formula_inventory(sheet, expected):
    actual = {}
    for coordinate, expected_formula in expected.items():
        cell = sheet[coordinate]
        actual_formula = formula_text(cell.value) if cell.data_type == "f" else None
        if actual_formula != expected_formula:
            raise ValueError(
                f"{coordinate}: expected {expected_formula!r}, got {actual_formula!r}"
            )
        actual[coordinate] = actual_formula
    return actual


check("extreme styled cell inflates the rectangular worksheet bounds (negative control)",
      formula_bound_ws.max_row == 1_048_576 and formula_bound_ws.max_column == 16_384,
      (formula_bound_ws.max_row, formula_bound_ws.max_column))
original_iter_rows = formula_bound_ws.iter_rows
formula_bound_ws.iter_rows = lambda *args, **kwargs: (_ for _ in ()).throw(
    RuntimeError("unbounded iter_rows must not run")
)
try:
    bounded_formulas = expected_formula_inventory(formula_bound_ws, {"D2": "=1+1"})
finally:
    formula_bound_ws.iter_rows = original_iter_rows
check("formula postcheck uses bounded public coordinate lookups",
      bounded_formulas == {"D2": "=1+1"}, bounded_formulas)

dimension_contract_wb = openpyxl.Workbook()
dimension_contract_ws = dimension_contract_wb.active
dimension_contract_ws.title = "Expected"
dimension_contract_ws["A1"] = "header"
dimension_contract_ws["C3"] = "tail"
expected_dimension_contract = {"Expected": "A1:C3"}
try:
    require(
        dimension_contract_ws.dimensions == expected_dimension_contract["Expected"],
        "unexpected used range",
    )
    exact_dimension_passed = True
except ValueError:
    exact_dimension_passed = False
dimension_contract_ws["D4"] = "unintended"
try:
    require(
        dimension_contract_ws.dimensions == expected_dimension_contract["Expected"],
        "unexpected used range",
    )
    stale_dimension_rejected = False
except ValueError:
    stale_dimension_rejected = True
check("postcheck accepts the exact task-declared worksheet range", exact_dimension_passed)
check("postcheck rejects unintended cells outside the task-declared range",
      stale_dimension_rejected, dimension_contract_ws.dimensions)
formula_bound_reopened.close()


# ---- edit.md snippet: extension detection is prefix-independent ------------------
X14_URI = b"http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
MC_URI = b"http://schemas.openxmlformats.org/markup-compatibility/2006"
EXTENSION_MARKERS = {
    "extLst": b"extLst",
    "x14 namespace": X14_URI,
    "markup compatibility": MC_URI,
}

def markers_in(data):
    return {label for label, marker in EXTENSION_MARKERS.items() if marker in data}

custom_prefix_sheet = (
    b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
    b'xmlns:sx="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main" '
    b'xmlns:ooo="http://schemas.openxmlformats.org/markup-compatibility/2006">'
    b'<ooo:AlternateContent><sx:extLst/></ooo:AlternateContent></worksheet>'
)
found = markers_in(custom_prefix_sheet)
check("namespace markers detect custom-prefix x14 extensions", "x14 namespace" in found, found)
check("namespace markers detect custom-prefix markup compatibility", "markup compatibility" in found, found)
check("local-name marker detects any-prefix extLst", "extLst" in found, found)
legacy_prefix_markers = {b"x14:", b"mc:AlternateContent"}
check("prefix markers are provably blind to custom prefixes (negative control)",
      not any(marker in custom_prefix_sheet for marker in legacy_prefix_markers))

# ---- formatting.md guard: header-only sheets skip conditional formatting ---------
def last_populated_row(sheet, *, first_data_row=2, min_col=1, max_col=6):
    populated_rows = (
        cell.row for cell in sheet._cells.values()
        if first_data_row <= cell.row
        and min_col <= cell.column <= max_col
        and cell.value is not None
    )
    return max(populated_rows, default=first_data_row - 1)


def add_demo_formatting(sheet):
    last = last_populated_row(sheet)
    if last < 2:
        return 0
    sheet.conditional_formatting.add(
        f"D2:D{last}", CellIsRule(operator="lessThan", formula=["0"]),
    )
    sheet.conditional_formatting.add(
        f"A2:F{last}", FormulaRule(formula=["$D2<0"]),
    )
    sheet.conditional_formatting.add(
        f"C2:C{last}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       end_type="max", end_color="63BE7B"),
    )
    sheet.conditional_formatting.add(
        f"E2:E{last}", DataBarRule(start_type="min", end_type="max", color="638EC6"),
    )
    return 4


header_wb = openpyxl.Workbook()
header_ws = header_wb.active
header_ws.append(["A", "B", "C", "D", "E", "F"])
header_ws["F100"].number_format = "0.00"  # styled empty cell inflates max_row
try:
    header_ws.conditional_formatting.add(
        "D2:D1", CellIsRule(operator="lessThan", formula=["0"]),
    )
    inverted_range_rejected = False
except (TypeError, ValueError):
    inverted_range_rejected = True
check("unguarded header-only range is rejected (negative control)", inverted_range_rejected)
check("style-only ghost row inflates max_row (negative control)", header_ws.max_row == 100)
check("populated-row scan ignores a style-only ghost row", last_populated_row(header_ws) == 1)
check("header-only guard skips all four formatting rules", add_demo_formatting(header_ws) == 0)
header_wb.save("header-only-formatting.xlsx")
header_reopened = openpyxl.load_workbook("header-only-formatting.xlsx")
check("header-only workbook saves and reopens with no conditional formatting",
      len(header_reopened.active.conditional_formatting) == 0)

data_wb = openpyxl.Workbook()
data_ws = data_wb.active
data_ws.append(["A", "B", "C", "D", "E", "F"])
data_ws.append([1, 2, 3, -1, 5, 6])
data_ws["F100"].number_format = "0.00"
check("data rows receive all four formatting rules", add_demo_formatting(data_ws) == 4)
data_ranges = {str(item.sqref) for item in data_ws.conditional_formatting}
check("conditional formatting stops at the last populated row despite ghost styles",
      data_ranges == {"D2", "A2:F2", "C2", "E2"}, data_ranges)
data_wb.save("data-formatting.xlsx")
data_reopened = openpyxl.load_workbook("data-formatting.xlsx")
check("all four formatting rules survive save/reopen",
      len(data_reopened.active.conditional_formatting) == 4)


# ---- read.md: implausible <dimension> is reset before streaming ------------------
dim_wb = openpyxl.Workbook()
dim_ws = dim_wb.active
dim_ws.append(["h1", "h2"])
dim_ws.append([1, 2])
dim_ws.append([3, 4])
dim_ws["C4"] = "=SUM(A2:B3)"
dim_wb.save("dimension.xlsx")
# Corrupt the sheet's dimension metadata the way non-Excel producers do.
import zipfile as dim_zip
with dim_zip.ZipFile("dimension.xlsx") as archive:
    members = {name: archive.read(name) for name in archive.namelist()}
members["xl/worksheets/sheet1.xml"] = members["xl/worksheets/sheet1.xml"].replace(
    b"<dimension ref=\"A1:C4\"/>", b"<dimension ref=\"A1:B2\"/>"
)
with dim_zip.ZipFile("dimension.xlsx", "w") as archive:
    for name, data in members.items():
        archive.writestr(name, data)

with validated_xlsx_source("dimension.xlsx") as dimension_source:
    dim_value = openpyxl.load_workbook(
        dimension_source, read_only=True, data_only=True
    )
    try:
        dim_ws_ro = dim_value.active
        check("plausible but truncated dimension limits streaming (negative control)",
              dim_ws_ro.calculate_dimension() == "A1:B2" and dim_ws_ro.max_row == 2,
              (dim_ws_ro.calculate_dimension(), dim_ws_ro.max_row))
        original_dimension_iter_rows = dim_ws_ro.iter_rows
        dim_ws_ro.iter_rows = lambda *args, **kwargs: (_ for _ in ()).throw(
            RuntimeError("XML discovery must not expand a worksheet rectangle")
        )
        dimension_source.seek(0)
        with zipfile.ZipFile(dimension_source) as archive:
            dimension_profile = worksheet_xml_profile(
                archive, worksheet_part(archive, dim_ws_ro.title)
            )
        dim_ws_ro.iter_rows = original_dimension_iter_rows
        dim_ws_ro.reset_dimensions()
        dimension_rows = list(bounded_sample_rows(dim_ws_ro, dimension_profile))
    finally:
        dim_value.close()
check("sparse XML profile ignores truncated dimension and retains uncached formulas",
      dimension_profile["extent"] == "A1:C4"
      and dimension_profile["first_populated_row"] == 1
      and dimension_profile["formula_count"] == 1
      and dimension_profile["missing_formula_samples"][0][0] == "C4",
      dimension_profile)
check("bounded sampling uses the XML extent after resetting producer metadata",
      len(dimension_rows) == 4
      and dimension_rows[2][:2] == (3, 4),
      dimension_rows)

missing_dimension_wb = openpyxl.Workbook()
missing_dimension_ws = missing_dimension_wb.active
missing_dimension_ws.append(["Region", "Units"])
missing_dimension_ws.append(["EU", 120])
missing_dimension_ws.append(["US", 80])
missing_dimension_wb.save("missing-dimension.xlsx")
with zipfile.ZipFile("missing-dimension.xlsx") as archive:
    missing_dimension_members = {
        name: archive.read(name) for name in archive.namelist()
    }
missing_dimension_xml = missing_dimension_members["xl/worksheets/sheet1.xml"]
missing_dimension_tag = b'<dimension ref="A1:B3"/>'
check("missing-dimension fixture starts with the expected producer metadata",
      missing_dimension_tag in missing_dimension_xml)
missing_dimension_members["xl/worksheets/sheet1.xml"] = (
    missing_dimension_xml.replace(missing_dimension_tag, b"")
)
with zipfile.ZipFile("missing-dimension.xlsx", "w") as archive:
    for name, data in missing_dimension_members.items():
        archive.writestr(name, data)

with validated_xlsx_source("missing-dimension.xlsx") as missing_dimension_source:
    missing_dimension_value_wb = openpyxl.load_workbook(
        missing_dimension_source, read_only=True, data_only=True
    )
    try:
        missing_dimension_value_ws = missing_dimension_value_wb.active
        forced_dimension_scans = []
        original_forced_dimension_scan = missing_dimension_value_ws._calculate_dimension

        def reject_forced_dimension_scan():
            forced_dimension_scans.append(True)
            raise RuntimeError("read route must not force a worksheet dimension scan")

        missing_dimension_value_ws._calculate_dimension = reject_forced_dimension_scan
        missing_dimension_source.seek(0)
        with zipfile.ZipFile(missing_dimension_source) as archive:
            missing_dimension_profile = worksheet_xml_profile(
                archive, worksheet_part(archive, missing_dimension_value_ws.title)
            )
        missing_dimension_declared = worksheet_declared_dimension(
            missing_dimension_value_ws
        )
        missing_dimension_value_ws.reset_dimensions()
        missing_dimension_rows = list(bounded_sample_rows(
            missing_dimension_value_ws, missing_dimension_profile
        ))
        missing_dimension_value_ws._calculate_dimension = original_forced_dimension_scan
    finally:
        missing_dimension_value_wb.close()
check("missing dimension remains unsized before raw XML routing",
      missing_dimension_declared is None
      and missing_dimension_profile["extent"] == "A1:B3",
      (missing_dimension_declared, missing_dimension_profile))
check("missing dimension route samples from XML bounds without a forced scan",
      forced_dimension_scans == []
      and missing_dimension_rows == [
          ("Region", "Units"), ("EU", 120), ("US", 80)
      ],
      (forced_dimension_scans, missing_dimension_rows))

csv_formula_wb = openpyxl.load_workbook("dimension.xlsx", read_only=True, data_only=False)
csv_value_wb = openpyxl.load_workbook("dimension.xlsx", read_only=True, data_only=True)
csv_formula_ws, csv_value_ws = csv_formula_wb.active, csv_value_wb.active
check("both CSV source streams initially trust the truncated dimension (negative control)",
      csv_formula_ws.calculate_dimension() == "A1:B2"
      and csv_value_ws.calculate_dimension() == "A1:B2")
csv_formula_ws.reset_dimensions()
csv_value_ws.reset_dimensions()
csv_stream_rows = list(zip(csv_formula_ws.iter_rows(), csv_value_ws.iter_rows()))
check("CSV export resets both paired streams before iterating",
      len(csv_stream_rows) == 4
      and len(csv_stream_rows[-1][0]) == 3
      and len(csv_stream_rows[-1][1]) == 3
      and csv_stream_rows[-1][0][2].data_type == "f"
      and csv_stream_rows[-1][1][2].value is None,
      [(len(formula_row), len(value_row)) for formula_row, value_row in csv_stream_rows])
csv_formula_wb.close()
csv_value_wb.close()

offset_wb = openpyxl.Workbook()
offset_ws = offset_wb.active
offset_ws["A7"], offset_ws["B7"] = "Region", "Units"
offset_ws["A8"], offset_ws["B8"] = "EU", 120
offset_ws["A2"].number_format = "0.00"  # styled but empty: not part of the data range
offset_ws["C8"] = "=SUM(B8)"            # uncached formula: remains part of the range
offset_ws["XFD1048576"].number_format = "0.00"  # extreme style-only physical cell
offset_wb.save("leading-blank-rows.xlsx")
with validated_xlsx_source("leading-blank-rows.xlsx") as offset_source:
    offset_value_wb = openpyxl.load_workbook(
        offset_source, read_only=True, data_only=True,
    )
    try:
        offset_value_ws = offset_value_wb.active
        original_offset_iter_rows = offset_value_ws.iter_rows
        offset_value_ws.iter_rows = lambda *args, **kwargs: (_ for _ in ()).throw(
            RuntimeError("XML profile must not call iter_rows")
        )
        offset_source.seek(0)
        with zipfile.ZipFile(offset_source) as archive:
            offset_profile = worksheet_xml_profile(
                archive, worksheet_part(archive, offset_value_ws.title)
            )
        offset_value_ws.iter_rows = original_offset_iter_rows
        offset_value_ws.reset_dimensions()
        offset_rows = bounded_sample_rows(offset_value_ws, offset_profile)
        offset_header = next(offset_rows, None)
        offset_sample = next(offset_rows, None)
    finally:
        offset_value_wb.close()
check("styled empty cells do not move the formula-preserving logical range",
      (offset_profile["extent"], offset_profile["first_populated_row"])
      == ("A7:C8", 7), offset_profile)
check("header sampling skips six leading blank rows",
      offset_header[:2] == ("Region", "Units")
      and offset_sample[:2] == ("EU", 120),
      (offset_header, offset_sample))

far_wb = openpyxl.Workbook()
far_ws = far_wb.active
far_ws["A1"] = "near"
far_ws["XFD1048576"] = "far"
far_wb.save("far-logical-cells.xlsx")
with validated_xlsx_source("far-logical-cells.xlsx") as far_source:
    far_value_wb = openpyxl.load_workbook(far_source, read_only=True, data_only=True)
    try:
        far_source.seek(0)
        with zipfile.ZipFile(far_source) as archive:
            far_profile = worksheet_xml_profile(
                archive, worksheet_part(archive, far_value_wb.active.title)
            )
        far_iter_calls = []
        far_value_wb.active.iter_rows = (
            lambda *args, **kwargs: far_iter_calls.append((args, kwargs)) or iter(())
        )
        try:
            bounded_sample_rows(far_value_wb.active, far_profile, max_cells=1_000)
            far_budget_rejected = False
        except ValueError as error:
            far_budget_rejected = "sample rectangle is too large" in str(error)
    finally:
        far_value_wb.close()
check("real far-apart values hit the rectangle budget before iter_rows",
      far_budget_rejected and far_iter_calls == [], (far_profile, far_iter_calls))

empty_dimension_wb = openpyxl.Workbook()
empty_dimension_wb.save("empty-dimension.xlsx")
with validated_xlsx_source("empty-dimension.xlsx") as empty_source:
    with zipfile.ZipFile(empty_source) as archive:
        empty_profile = worksheet_xml_profile(
            archive, worksheet_part(archive, "Sheet")
        )
check("dimension scan handles an actually empty worksheet",
      empty_profile["extent"] == "A1:A1"
      and empty_profile["first_populated_row"] is None)


print("\n" + ("ALL XLSX FIXTURES PASSED" if not failures else f"{len(failures)} FAILURES: {failures}"))
sys.exit(0 if not failures else 1)
